import csv
import openpyxl
import pandas as pd



################# 新規作成#############################
######################################################
wb = openpyxl.Workbook()
ws = wb["Sheet"]

###################### データ#########################
######################################################
## カラム= class,numの設定
## classはA,B,Cのいずれか
## numは1-12のいずれか
df = pd.DataFrame([
        ['A',5], 
        ['B',9], 
        ['C',3], 
        ['C',5], 
        ['C',2], 
        ['C',10], 
        ['B',12],
        ['A',6],
        ['B',7],
        ['C',1], 
        ],columns=['class','num'])

## カラム名とカラム数の取得
temp_column_name = []
for column_name in df:
        temp_column_name.append(column_name)
LENGTH = len(df[temp_column_name[0]])

## 特定のカラム名(class)の値の取得(重複なし)
column_kind = []
for i in range(LENGTH):
        column_kind.append(df[temp_column_name[0]][i])
column_kind = list(dict.fromkeys(column_kind))
LENGTH_COLUMN_KIND = len(column_kind)

##################### 番号(1-12)の表示####################
#########################################################
START_COLUMU = 2
for alphabet,num in zip (range(66,91),range(1,13)):
    ws[f"{chr(alphabet)}{START_COLUMU}"].value = num


################ カラム>>>classの表示#####################
#########################################################
for i in range(LENGTH_COLUMN_KIND):
    ws[f"A{3 + i}"].value = column_kind[i]


################ カラム>>>numの表示#####################
#########################################################
## 該当のnumがある場合1を出力
for i in range(LENGTH_COLUMN_KIND):
    for class_,num in zip(df[temp_column_name[0]],df[temp_column_name[1]]):
        if class_ == column_kind[i]:
            ws[f"{chr(65 + num)}{3 + i}"].value = 1


wb.save('auto.xlsx')
wb.close()
