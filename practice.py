import time as tm
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

################# 元ファイルの読み込み#############################
#################################################################
datafile = './original.xlsx'
sheetname = 'Sheet'

wb = load_workbook(filename=datafile)
ws_in = wb[sheetname]
df_ori = pd.DataFrame(ws_in.values)

# 欠損値NaNを0に置換
df_ori = df_ori.fillna(0)
print(df_ori)


############### 追加ファイルの読み込み#############################
#################################################################
datafile_add = './auto.xlsx'
sheetname = 'Sheet'

wb_add = load_workbook(filename=datafile_add)
ws_in_add = wb_add[sheetname]
df_add = pd.DataFrame(ws_in_add.values)

# 欠損値NaNを0に置換
df_add = df_add.fillna(0)
print(df_add)

############### データの解析#############################
#################################################################
# クラスが共通かの判定
ori_column = []
for ori_col in df_ori:
        ori_column.append(ori_col)

add_column = []
for add_col in df_add:
        add_column.append(add_col)

if ori_column == add_column:
        print(0)
else:
        print(1)





wb.save(datafile)
wb.close()

wb_add.save(datafile_add)
wb_add.close()

